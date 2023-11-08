import openpyxl
import smtplib
from email.message import EmailMessage

# Global variables to store data
attendance_file = "attendance.xlsx"
threshold = 0.8  # Threshold for attendance percentage
email_config = {
    "smtp_server": "smtp.example.com",
    "smtp_port": 587,
    "sender_email": "your_email@example.com",
    "sender_password": "your_password",
    "subject": "Attendance Notification",
    "message": "You have been marked absent for today's class.",
}

def load_attendance_data():
    try:
        workbook = openpyxl.load_workbook(attendance_file)
        sheet = workbook.active
        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        return data

    except Exception as e:
        print(f"Error loading attendance data: {e}")

def save_attendance_data(data):
    try:
        workbook = openpyxl.load_workbook(attendance_file)
        sheet = workbook.active

        for i, row in enumerate(data):
            for j, value in enumerate(row):
                sheet.cell(row=i + 2, column=j + 1, value=value)

        workbook.save(attendance_file)

    except Exception as e:
        print(f"Error saving attendance data: {e}")

def send_email(recipient_email):
    try:
        msg = EmailMessage()
        msg.set_content(email_config["message"])
        msg["Subject"] = email_config["subject"]
        msg["From"] = email_config["sender_email"]
        msg["To"] = recipient_email

        server = smtplib.SMTP(email_config["smtp_server"], email_config["smtp_port"])
        server.starttls()
        server.login(email_config["sender_email"], email_config["sender_password"])
        server.send_message(msg)
        server.quit()

        print(f"Email sent to {recipient_email}")

    except Exception as e:
        print(f"Error sending email: {e}")

def track_attendance():
    attendance_data = load_attendance_data()
    total_students = len(attendance_data)
    present_students = 0

    for student in attendance_data:
        if student[1] == "Present":
            present_students += 1
        elif student[1] == "Absent":
            send_email(student[0])

    attendance_percentage = present_students / total_students

    if attendance_percentage >= threshold:
        print("Attendance is above the threshold.")
    else:
        print("Attendance is below the threshold. Notifications sent to absentees.")

if __name__ == "__main__":
    track_attendance()
